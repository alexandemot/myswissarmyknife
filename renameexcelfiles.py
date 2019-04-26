
import openpyxl
from openpyxl import workbook
import datetime


time = (datetime.datetime.now()).strftime("%Y%m%d%H%M%S")

workbook = openpyxl.load_workbook('sample-excel.xlsx')
sheetnames = workbook.sheetnames

sheet = workbook.active

maxrow = sheet.max_row
maxcolumn = sheet.max_column


'''
for i in range(1,maxrow+1):
	for j in range(1,maxcolumn+1):
		cell = sheet.cell(i,j)
		print(cell.value)
'''


for each in sheetnames:
	newname = each+'_'+time
	file = open("%s.txt" %newname, "w+")
	file.close()
