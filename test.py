
import openpyxl
from openpyxl import workbook
import datetime


time = (datetime.datetime.now()).strftime("%Y%m%d%H%M%S")

workbook = openpyxl.load_workbook('test.xlsx')
sheetnames = workbook.sheetnames


for each in sheetnames:
	newname = each+'_'+time
	file = open("%s.txt" %newname, "w+")
	file.close()

